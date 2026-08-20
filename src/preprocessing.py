# """
# preprocessing.py
# -----------------
# Loaders for the four datasets used in the assignment:

#   Classification:
#     - Linearly Separable (LS):  data/Group23/classification/LS_Group23/Class{1,2,3}.txt
#     - Non-linearly Separable (NLS): data/Group23/classification/NLS_Group23.txt

#   Regression:
#     - Univariate: data/Group23/regression/Univariate/Univariate.xlsx
#     - Bivariate:  data/Group23/regression/Bivariate/Bivariate.xlsx

# IMPORTANT
# ---------
# These loaders were written from the assignment description alone
# (the actual data files were not available at the time this code was
# written). The parsers below are deliberately defensive/flexible about
# whitespace vs comma delimiters and about an optional header line, but
# you MUST sanity-check the printed shapes (`--check-data` flag in
# main.py) against your real files the first time you run this, and
# adjust `load_nls_data()`'s header-parsing if your file's "number of
# examples per class" header line is formatted differently.
# """

# import numpy as np
# import os


# def _load_numeric_matrix(path):
#     """
#     Read a whitespace/comma separated numeric text file into a 2-D
#     numpy array, skipping any non-numeric header line(s) automatically.
#     """
#     rows = []
#     with open(path, "r") as f:
#         for line in f:
#             line = line.strip()
#             if not line:
#                 continue
#             parts = line.replace(",", " ").split()
#             try:
#                 rows.append([float(p) for p in parts])
#             except ValueError:
#                 # non-numeric line (header / comment) -> skip it
#                 continue
#     return np.array(rows)


# # ------------------------- classification -------------------------

# def load_ls_data(data_dir):
#     """
#     Loads the 3-class Linearly Separable dataset.
#     Expects: Class1.txt, Class2.txt, Class3.txt inside `data_dir`,
#     each row being "x1 x2" (or "x1,x2") for one sample.

#     Returns
#     -------
#     X : (N, 2) ndarray
#     y : (N,) ndarray of int class labels {1, 2, 3}
#     """
#     X_list, y_list = [], []
#     class_files = sorted(
#         f for f in os.listdir(data_dir)
#         if f.lower().startswith("class") and f.lower().endswith(".txt")
#     )
#     if not class_files:
#         raise FileNotFoundError(
#             f"No Class*.txt files found in {data_dir}. "
#             "Place Class1.txt, Class2.txt, Class3.txt there."
#         )

#     for fname in class_files:
#         # extract the trailing integer in e.g. "Class1.txt" -> 1
#         digits = "".join(ch for ch in fname if ch.isdigit())
#         label = int(digits) if digits else None
#         if label is None:
#             raise ValueError(f"Could not infer class label from filename {fname}")

#         mat = _load_numeric_matrix(os.path.join(data_dir, fname))
#         X_list.append(mat[:, :2])
#         y_list.append(np.full(mat.shape[0], label))

#     X = np.vstack(X_list)
#     y = np.concatenate(y_list)
#     return X, y


# def load_nls_data(path, class_labels=(1, 2)):
#     """
#     Loads the Non-linearly Separable dataset (2 or 3 classes) from a
#     single file. Per the assignment: "The number of examples in each
#     class and their order is given at the beginning of each file."

#     This function first tries to interpret the FIRST non-empty line as
#     a whitespace/comma separated list of per-class counts (e.g. "500
#     500" or "500 500 500"); the remaining lines are then split into
#     class blocks IN THAT ORDER, matching `class_labels` positionally.

#     If the first line does NOT look like a small list of integer
#     counts (e.g. it's actually numeric x1/x2 data), the whole file is
#     instead assumed to already contain a label column (x1, x2, label)
#     and is loaded directly.

#     Returns
#     -------
#     X : (N, 2) ndarray
#     y : (N,) ndarray of int class labels
#     """
#     with open(path, "r") as f:
#         lines = [ln.strip() for ln in f if ln.strip()]

#     header_parts = lines[0].replace(",", " ").split()
#     is_count_header = (
#         1 <= len(header_parts) <= 3
#         and all(p.lstrip("-").isdigit() for p in header_parts)
#     )

#     if is_count_header:
#         counts = [int(p) for p in header_parts]
#         data_lines = lines[1:]
#         rows = []
#         for ln in data_lines:
#             parts = ln.replace(",", " ").split()
#             rows.append([float(p) for p in parts])
#         data = np.array(rows)

#         X_list, y_list = [], []
#         start = 0
#         for count, label in zip(counts, class_labels):
#             block = data[start:start + count]
#             X_list.append(block[:, :2])
#             y_list.append(np.full(block.shape[0], label))
#             start += count
#         X = np.vstack(X_list)
#         y = np.concatenate(y_list)
#         return X, y

#     # fallback: assume every row is "x1 x2 label"
#     rows = []
#     for ln in lines:
#         parts = ln.replace(",", " ").split()
#         rows.append([float(p) for p in parts])
#     data = np.array(rows)
#     X = data[:, :2]
#     y = data[:, 2].astype(int)
#     return X, y


# # ------------------------- regression -------------------------

# def _load_xlsx_as_array(path):
#     """Loads an .xlsx file's first sheet into a numeric numpy array
#     using openpyxl (no pandas dependency required)."""
#     from openpyxl import load_workbook
#     wb = load_workbook(path, read_only=True, data_only=True)
#     ws = wb.active
#     rows = []
#     for row in ws.iter_rows(values_only=True):
#         if row is None:
#             continue
#         if any(v is None for v in row):
#             continue
#         try:
#             rows.append([float(v) for v in row])
#         except (TypeError, ValueError):
#             continue  # header row
#     return np.array(rows)


# def load_univariate_regression(path):
#     """Returns X (N,1), y (N,) for the univariate regression dataset.
#     Assumes columns are (x, y) in that order."""
#     data = _load_xlsx_as_array(path)
#     X = data[:, [0]]
#     y = data[:, 1]
#     return X, y


# def load_bivariate_regression(path):
#     """Returns X (N,2), y (N,) for the bivariate regression dataset.
#     Assumes columns are (x1, x2, y) in that order."""
#     data = _load_xlsx_as_array(path)
#     X = data[:, [0, 1]]
#     y = data[:, 2]
#     return X, y
"""
preprocessing.py
-----------------
Loaders for the four datasets used in the assignment:

  Classification:
    - Linearly Separable (LS):  data/Group23/classification/LS_Group23/Class{1,2,3}.txt
    - Non-linearly Separable (NLS): data/Group23/classification/NLS_Group23.txt

  Regression:
    - Univariate: data/Group23/regression/Univariate/23.csv
    - Bivariate:  data/Group23/regression/Bivariate/23.csv
"""
"""
preprocessing.py
-----------------
Loaders for classification and regression datasets.
"""

import os
import re
import numpy as np


def _resolve_path(path):
    """
    If the specified path does not exist, search its directory
    for any valid data file (.csv, .xlsx, .txt).
    """
    if os.path.exists(path):
        return path
    
    # Try directory of the path or the path itself if it's a directory
    dirname = path if os.path.isdir(path) else os.path.dirname(path)
    
    if os.path.exists(dirname):
        for f in os.listdir(dirname):
            if f.lower().endswith(('.csv', '.xlsx', '.txt')):
                return os.path.join(dirname, f)
                
    return path


def _load_numeric_matrix(path):
    """Read a whitespace or comma-separated numeric file into a NumPy array."""
    path = _resolve_path(path)
    rows = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.replace(",", " ").split()
            try:
                rows.append([float(p) for p in parts])
            except ValueError:
                continue
    return np.array(rows)


# ------------------------- Classification -------------------------

def load_ls_data(data_dir):
    """Loads the 3-class Linearly Separable dataset."""
    data_dir = _resolve_path(data_dir)
    X_list, y_list = [], []
    class_files = sorted(
        f for f in os.listdir(data_dir)
        if f.lower().startswith("class") and f.lower().endswith(".txt")
    )
    if not class_files:
        raise FileNotFoundError(
            f"No Class*.txt files found in {data_dir}."
        )

    for fname in class_files:
        digits = "".join(ch for ch in fname if ch.isdigit())
        label = int(digits) if digits else None
        if label is None:
            raise ValueError(f"Could not infer class label from filename {fname}")

        mat = _load_numeric_matrix(os.path.join(data_dir, fname))
        X_list.append(mat[:, :2])
        y_list.append(np.full(mat.shape[0], label))

    X = np.vstack(X_list)
    y = np.concatenate(y_list)
    return X, y


def load_nls_data(path, class_labels=(1, 2)):
    """Loads the Non-linearly Separable dataset from a single file."""
    path = _resolve_path(path)
    with open(path, "r") as f:
        lines = [ln.strip() for ln in f if ln.strip()]

    header_lines = []
    numeric_rows = []

    for line in lines:
        parts = line.replace(",", " ").split()
        try:
            row = [float(p) for p in parts]
            numeric_rows.append(row)
        except ValueError:
            header_lines.append(line)

    if not numeric_rows:
        raise ValueError(f"No numeric data rows found in {path}")

    data = np.array(numeric_rows)

    if data.shape[1] >= 3:
        X = data[:, :2]
        y = data[:, -1].astype(int)
        return X, y

    counts = []
    if header_lines:
        found_nums = [int(num) for num in re.findall(r'\d+', " ".join(header_lines))]
        if len(found_nums) >= len(class_labels):
            counts = found_nums[:len(class_labels)]

    if not counts or sum(counts) != len(data):
        n_classes = len(class_labels)
        base_count = len(data) // n_classes
        counts = [base_count] * n_classes

    X_list, y_list = [], []
    start = 0
    for count, label in zip(counts, class_labels):
        block = data[start : start + count]
        X_list.append(block[:, :2])
        y_list.append(np.full(len(block), label))
        start += count

    X = np.vstack(X_list)
    y = np.concatenate(y_list)
    return X, y


# ------------------------- Regression -------------------------

def _load_xlsx_as_array(path):
    """Loads .csv, .txt, or .xlsx files into a numeric numpy array."""
    resolved_path = _resolve_path(path)
    
    if not os.path.exists(resolved_path):
        raise FileNotFoundError(f"Could not find regression data file at {path} or {resolved_path}")

    if resolved_path.endswith('.csv') or resolved_path.endswith('.txt'):
        return _load_numeric_matrix(resolved_path)

    from openpyxl import load_workbook
    wb = load_workbook(resolved_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None or any(v is None for v in row):
            continue
        try:
            rows.append([float(v) for v in row])
        except (TypeError, ValueError):
            continue
    return np.array(rows)


def load_univariate_regression(path):
    """Returns X (N,1), y (N,) for univariate regression."""
    data = _load_xlsx_as_array(path)
    X = data[:, [0]]
    y = data[:, 1]
    return X, y


def load_bivariate_regression(path):
    """Returns X (N,2), y (N,) for bivariate regression."""
    data = _load_xlsx_as_array(path)
    X = data[:, [0, 1]]
    y = data[:, 2]
    return X, y